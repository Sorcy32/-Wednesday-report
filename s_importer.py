from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl

global_types = {} #принадлежсноть адресов к типам
Errors = 0
Curr_row=0
all_types=[]


def import_global_types():
    '''
    #Получаем принадлежности объектов к типам объектов
    '''
    wb = load_workbook('sdata\ObjectTypesGlobal.xlsx')
    ws = wb["Names"]        #Указываем из какой страницы брать data
    for row in ws.values:
        global_types[row[1]] = row[3]
    langs = list()
    for lang in global_types.values():
        if type(lang) == list:
            langs += lang
        else:
            langs.append(lang)
    global all_types
    all_types =(set(langs))
    print('Всего типов найдено: ', all_types)


def import_input(): #Грузим входной файл
    wb=load_workbook('input.xlsx')
    ws = wb.active # активная книга
    x = ws.max_row
    wbe = Workbook() #Книга для сохранения
    print('\n Всего строк для обработки: ', x)    
    for typ in all_types:
        wse = wbe.create_sheet(typ)
    fst=True    
    for row in ws.values:
        if fst==True:
            fst=False
        else:
            global Curr_row
            Curr_row +=1
            try:
                save=wbe[(global_types.get(row[3]))]
                save.append(row)
                save=wbe['Sheet']
                save.append(row)
            except:
                global Errors
                Errors+=1
                print('Ошибка номер: ', Errors,
                      'В строке № ', Curr_row, ' входного файла' )                
    try:
        del wbe['Принадлежность']
    except:
        print('Ошибка удаления временной таблицы. Не критично.')
    wbe.save('exp.xlsx')
    print('Сохранено в exp.xlsx')

if __name__ == '__main__':
    import_global_types()
    import_input()




